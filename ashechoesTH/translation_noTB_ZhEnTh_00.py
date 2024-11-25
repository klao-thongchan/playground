import pandas as pd
import anthropic
from openpyxl.styles import PatternFill
import re

class TranslationSystem:
    def __init__(self, api_key: str):
        self.client = anthropic.Client(api_key=api_key)
        
    def translate_file(self, input_file: str, output_file: str) -> None:
        print(f"Processing {input_file}...")
        
        # Read input file
        df = pd.read_excel(input_file, header=None)
        
        # Define fill colors
        red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        
        # Process each row
        total_rows = len(df)
        for idx, row in df.iterrows():
            chinese_text = str(row.iloc[1]).strip()  # Column B
            english_text = str(row.iloc[2]).strip()  # Column C
            existing_thai = str(row.iloc[3]).strip() if len(row.iloc) > 3 else ""  # Column D
            
            # Skip if no Chinese text or if valid translation exists
            if pd.isna(chinese_text) or not chinese_text or (existing_thai and existing_thai != "#N/A"):
                continue
            
            # Prepare context for translation
            sources = f"Chinese: {chinese_text}"
            if english_text and not pd.isna(english_text):
                sources += f"\nEnglish: {english_text}"
            
            # Create prompt for translation and spell check
            prompt = f"""Translate the following text to Thai. Maintain any symbols (< > etc.) from the source text.

Source text:
{sources}

If Thai translation exists but seems incorrect according to sources, reply with INCORRECT: followed by the correct translation.
If Thai text needs spelling fixes, reply with SPELLING: followed by the corrected text.
Otherwise, provide only the Thai translation."""

            try:
                # Get translation from Claude
                message = self.client.messages.create(
                    model="claude-3-5-haiku-latest",
                    max_tokens=1000,
                    temperature=0.3,
                    messages=[{"role": "user", "content": prompt}]
                )
                
                response = message.content[0].text.strip()
                
                # Process the response
                if response.startswith("INCORRECT:"):
                    df.iloc[idx, 3] = response[10:].strip()
                    df.at[idx, 'color'] = 'red'
                elif response.startswith("SPELLING:"):
                    df.iloc[idx, 3] = response[9:].strip()
                    df.at[idx, 'color'] = 'yellow'
                else:
                    df.iloc[idx, 3] = response
                
            except Exception as e:
                print(f"Error processing row {idx + 1}: {str(e)}")
                df.iloc[idx, 3] = f"ERROR: {str(e)}"
            
            if (idx + 1) % 10 == 0:
                print(f"Processed {idx + 1}/{total_rows} rows")
        
        # Save with color formatting
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        df.to_excel(writer, index=False, header=False)
        
        # Apply colors
        worksheet = writer.sheets['Sheet1']
        for idx, row in df.iterrows():
            if df.at[idx, 'color'] == 'red':
                worksheet.cell(idx + 2, 4).fill = red_fill
            elif df.at[idx, 'color'] == 'yellow':
                worksheet.cell(idx + 2, 4).fill = yellow_fill
        
        writer.close()
        print(f"Processing completed. Output saved to {output_file}")

def main():
    API_KEY = "your-api-key"  # Replace with your API key
    INPUT_FILE = "input.xlsx"
    OUTPUT_FILE = "output.xlsx"
    
    translator = TranslationSystem(API_KEY)
    translator.translate_file(INPUT_FILE, OUTPUT_FILE)

if __name__ == "__main__":
    main()